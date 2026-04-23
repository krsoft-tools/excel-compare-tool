from fastapi import FastAPI, UploadFile, File, Form
from fastapi.responses import HTMLResponse, StreamingResponse
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

app = FastAPI()

@app.get("/", response_class=HTMLResponse)
def home():
    return """
    <html>
    <body style="font-family: Arial; text-align:center; padding:50px;">

    <h1>Find changes in Excel instantly</h1>

    <form action="/export" method="post" enctype="multipart/form-data">
        <input type="file" name="file1"><br><br>
        <input type="file" name="file2"><br><br>
        <button type="submit">Compare</button>
    </form>

    </body>
    </html>
    """

# =========================
# NORMALIZATION
# =========================
def normalize_value(val):
    if pd.isna(val) or val == "":
        return ""
    try:
        f = float(val)
        if f.is_integer():
            return str(int(f))
        return str(f)
    except:
        return str(val).strip()


# =========================
# CORE DATA PREP
# =========================
async def prepare_data(file1, file2, key):
    df1 = pd.read_excel(file1.file, engine="openpyxl").fillna("")
    df2 = pd.read_excel(file2.file, engine="openpyxl").fillna("")

    df1.columns = df1.columns.str.strip()
    df2.columns = df2.columns.str.strip()

    for col in df1.columns:
        if 'date' in col.lower():
            df1[col] = pd.to_datetime(df1[col], errors='coerce')

    for col in df2.columns:
        if 'date' in col.lower():
            df2[col] = pd.to_datetime(df2[col], errors='coerce')

    # remove junk
    df1 = df1.loc[:, ~df1.columns.str.contains('^Unnamed')]
    df2 = df2.loc[:, ~df2.columns.str.contains('^Unnamed')]

    if key not in df1.columns or key not in df2.columns:
        return None, None, None, None, f"Column '{key}' not found"

    # remove empty rows
    df1 = df1[df1[key].notna()]
    df2 = df2[df2[key].notna()]

    df1 = df1[df1[key].astype(str).str.strip() != ""]
    df2 = df2[df2[key].astype(str).str.strip() != ""]

    # normalize
    df1[key] = df1[key].apply(normalize_value)
    df2[key] = df2[key].apply(normalize_value)

    for col in df1.columns:
        df1[col] = df1[col].apply(normalize_value)
    for col in df2.columns:
        df2[col] = df2[col].apply(normalize_value)

    # missing
    keys1 = set(df1[key])
    keys2 = set(df2[key])

    missing_in_file2 = sorted(list(keys1 - keys2))
    missing_in_file1 = sorted(list(keys2 - keys1))

    # merge
    df1 = df1.reset_index(drop=True)
    df2 = df2.reset_index(drop=True)
    
    merged = df1.merge(df2, on=key, how="outer", suffixes=("_file1", "_file2"))

    return df1, df2, merged, missing_in_file2, missing_in_file1, None


# =========================
# DIFFERENCE LOGIC (for UI)
# =========================
def extract_differences(df1, merged, key):
    differences = []

    for _, row in merged.iterrows():
        diff = {}

        for col in df1.columns:
            if col == key:
                continue

            col1 = f"{col}_file1"
            col2 = f"{col}_file2"

            if col1 in row and col2 in row:
                val1 = row[col1]
                val2 = row[col2]

                if val1 != val2:
                    diff[col] = {
                        "file1": val1,
                        "file2": val2
                    }

        if diff:
            differences.append({
                "id": row[key],
                "differences": diff
            })

    return differences


# =========================
# API (JSON)
# =========================
@app.post("/compare")
async def compare_files(
    file1: UploadFile = File(...),
    file2: UploadFile = File(...),
    key: str = Form(...)
):
    df1, df2, merged, missing_in_file2, missing_in_file1, error = await prepare_data(file1, file2, key)

    if error:
        return {"error": error}

    differences = extract_differences(df1, merged, key)

    return {
        "rows_compared": len(merged),
        "differences_found": len(differences),
        "differences": differences,
        "missing_in_file2": missing_in_file2,
        "missing_in_file1": missing_in_file1
    }


# =========================
# UI FORM
# =========================
@app.get("/ui", response_class=HTMLResponse)
def ui():
    return """
    <html>
    <head>
        <title>Excel Compare</title>
        <style>
            body {
                font-family: Arial;
                background: #f5f7fa;
                display: flex;
                justify-content: center;
                padding-top: 50px;
            }
            .card {
                background: white;
                padding: 30px;
                border-radius: 10px;
                box-shadow: 0 10px 30px rgba(0,0,0,0.1);
                width: 400px;
            }
            input, button {
                width: 100%;
                margin-bottom: 10px;
                padding: 10px;
            }
            button {
                background: #4CAF50;
                color: white;
                border: none;
                cursor: pointer;
            }
        </style>
    </head>
    <body>
        <div class="card">
            <h2>Compare Excel Files</h2>
            <form method="post" enctype="multipart/form-data">
                <input type="file" name="file1">
                <input type="file" name="file2">
                <input type="text" name="key" placeholder="Key column">

                <button formaction="/export">Compare & Download Excel</button>
            </form>
        </div>
    </body>
    </html>
    """


# =========================
# UI RESULT
# =========================
@app.post("/compare-ui", response_class=HTMLResponse)
async def compare_ui(file1: UploadFile = File(...), file2: UploadFile = File(...), key: str = Form(...)):
    df1, df2, merged, missing_in_file2, missing_in_file1, error = await prepare_data(file1, file2, key)

    if error:
        return f"<h3>{error}</h3>"

    differences = extract_differences(df1, merged, key)

    html = """
    <html>
    <head>
    <style>
        body { font-family: Arial; background: #f5f7fa; padding: 40px; }
        .container { max-width: 800px; margin: auto; }
        .card { background: white; padding: 20px; margin-bottom: 20px; border-radius: 10px; }
        table { width: 100%; border-collapse: collapse; }
        td, th { padding: 10px; border-bottom: 1px solid #eee; }
        .diff { background: #ffe5e5; }
        .missing { background: #fff2cc; padding: 10px; margin: 5px 0; }
    </style>
    </head>
    <body><div class="container">
    """

    html += f"<div class='card'><h2>Differences: {len(differences)}</h2></div>"

    html += "<div class='card'><table><tr><th>ID</th><th>Column</th><th>File1</th><th>File2</th></tr>"

    for d in differences:
        for col, vals in d["differences"].items():
            html += f"""
            <tr>
                <td>{d['id']}</td>
                <td>{col}</td>
                <td class='diff'>{vals['file1']}</td>
                <td class='diff'>{vals['file2']}</td>
            </tr>
            """

    html += "</table></div>"

    html += "<div class='card'><h3>Missing in file2</h3>"
    for m in missing_in_file2:
        html += f"<div class='missing'>{m}</div>"
    html += "</div>"

    html += "<div class='card'><h3>Missing in file1</h3>"
    for m in missing_in_file1:
        html += f"<div class='missing'>{m}</div>"
    html += "</div>"

    html += "</div></body></html>"

    return html


# =========================
# EXPORT WITH HIGHLIGHTS
# =========================
@app.post("/export")
async def export_excel(
    file1: UploadFile = File(...),
    file2: UploadFile = File(...)
):
    print("NEW EXPORT CODE LOADED")

    import difflib
    from io import BytesIO
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    # 👉 reset stream (KRITICKÉ)
    content1 = await file1.read()
    content2 = await file2.read()

    # 👉 load
    df1 = pd.read_excel(file1.file, engine="openpyxl").fillna("")
    df2 = pd.read_excel(file2.file, engine="openpyxl").fillna("")

    # 👉 clean columns
    df1.columns = df1.columns.str.strip()
    df2.columns = df2.columns.str.strip()

    df1 = df1.loc[:, ~df1.columns.str.contains('^Unnamed')]
    df2 = df2.loc[:, ~df2.columns.str.contains('^Unnamed')]

    # 👉 unify columns (DÔLEŽITÉ pre delete)
    all_cols = list(df1.columns.union(df2.columns))
    df1 = df1.reindex(columns=all_cols, fill_value="")
    df2 = df2.reindex(columns=all_cols, fill_value="")

    # 👉 datetime fix
    for col in df1.columns:
        if 'date' in col.lower():
            df1[col] = pd.to_datetime(df1[col], errors='coerce').dt.strftime('%Y-%m-%d')
    for col in df2.columns:
        if 'date' in col.lower():
            df2[col] = pd.to_datetime(df2[col], errors='coerce').dt.strftime('%Y-%m-%d')

    # 👉 normalize
    def normalize(val):
        if pd.isna(val) or val == "":
            return ""
        try:
            f = float(val)
            if f.is_integer():
                return str(int(f))
            return str(f)
        except:
            return str(val).strip()

    df1 = df1.apply(lambda col: col.map(normalize))
    df2 = df2.apply(lambda col: col.map(normalize))

    # 👉 rows to string
    def row_to_str(row):
        return "|".join([normalize(v) for v in row])

    rows1 = [row_to_str(r) for r in df1.values]
    rows2 = [row_to_str(r) for r in df2.values]

    matcher = difflib.SequenceMatcher(None, rows1, rows2)

    output_rows = []
    row_types = []

    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        print(tag, i1, i2, j1, j2)

        if tag == "equal":
            for i in range(i1, i2):
                output_rows.append(df2.iloc[i].tolist())
                row_types.append("equal")

        elif tag == "replace":
            len1 = i2 - i1
            len2 = j2 - j1

            # 👉 rovnaký počet = skutočný replace
            if len1 == len2:
                for k in range(len1):
                    output_rows.append(df2.iloc[j1 + k].tolist())
                    row_types.append("replace")
            else:
                # 👉 split na delete + insert
                for i in range(i1, i2):
                    output_rows.append(df1.iloc[i].tolist())
                    row_types.append("delete")

                for j in range(j1, j2):
                    output_rows.append(df2.iloc[j].tolist())
                    row_types.append("insert")

        elif tag == "insert":
            for j in range(j1, j2):
                output_rows.append(df2.iloc[j].tolist())
                row_types.append("insert")

        elif tag == "delete":
            for i in range(i1, i2):
                output_rows.append(df1.iloc[i].tolist())
                row_types.append("delete")

    result_df = pd.DataFrame(output_rows, columns=all_cols)

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        result_df.to_excel(writer, index=False, sheet_name="result")

    output.seek(0)

    wb = load_workbook(output)
    ws = wb["result"]

    orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    green = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    red = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")

    for idx, row in enumerate(ws.iter_rows(min_row=2), start=0):
        t = row_types[idx]

        if t == "insert":
            for c in row:
                c.fill = green

        elif t == "delete":
            for c in row:
                c.fill = red

        elif t == "replace":
            try:
                original = rows1[idx].split("|")
            except:
                original = [""] * len(row)

            for col_idx, c in enumerate(row):
                val1 = original[col_idx] if col_idx < len(original) else ""
                val2 = c.value if c.value is not None else ""

                if str(val1) != str(val2):
                    c.fill = orange

    final = BytesIO()
    wb.save(final)
    final.seek(0)

    return StreamingResponse(
        final,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=diff.xlsx"}
    )